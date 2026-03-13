from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Dict, List, Optional
import zipfile

import pandas as pd
import requests
from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).resolve().parent / "New_Industry Classifications.xlsx"
SOURCE_SHEET = "Tab1"
OUTPUT_SHEET = "FF49_Comparison"
SUMMARY_SHEET = "FF49_Agreement"
SIC49_URL = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/Siccodes49.zip"

DOMAIN_RULES: Dict[str, Dict[str, List[str]]] = {
    "Technology": {
        "very_strong": [
            "semiconductor", "chip", "processor", "software", "cybersecurity", "cloud",
        ],
        "strong": [
            "software", "saas", "internet", "data", "cloud", "semiconductor", "chip",
            "electronic", "electronics", "hardware", "telecom", "network", "cyber",
            "ai", "server", "processor", "memory", "storage", "digital", "it ",
        ],
        "medium": ["information", "platform", "infrastructure", "communications", "computing"],
        "weak": ["components", "systems", "automation", "services"],
    },
    "Financials": {
        "very_strong": ["banking", "insurance", "asset management", "investment services", "reit"],
        "strong": [
            "bank", "banking", "insurance", "finance", "financial", "credit", "lending",
            "investment", "broker", "asset management", "reit", "mortgage", "clearing",
        ],
        "medium": ["payment processing", "transaction processing", "depository", "underwriting"],
        "weak": ["capital", "portfolio", "transaction", "payment", "trust", "holdings"],
    },
    "Healthcare": {
        "very_strong": ["biopharmaceutical", "biotech", "medical devices", "hospital"],
        "strong": [
            "health", "healthcare", "biopharmaceutical", "pharma", "drug", "medical",
            "biotech", "hospital", "clinic", "dental", "therapeutic", "diagnostic",
        ],
        "medium": ["clinical", "life science", "laboratory", "genomics", "diagnostics"],
        "weak": ["care", "devices", "wellness"],
    },
    "Energy": {
        "very_strong": ["upstream", "exploration and production", "oil and gas"],
        "strong": [
            "oil", "gas", "petroleum", "fossil fuel", "coal", "energy", "drilling",
            "upstream", "exploration", "pipeline", "refining",
        ],
        "medium": ["midstream", "downstream", "offshore", "natural gas", "lng"],
        "weak": ["well", "field services"],
    },
    "Utilities": {
        "very_strong": ["electric utilities", "gas utilities", "water utilities"],
        "strong": ["utilities", "utility", "electric", "power", "water", "steam", "grid"],
        "medium": ["transmission", "distribution", "generation"],
        "weak": ["renewable", "municipal utility"],
    },
    "Real Estate": {
        "very_strong": ["real estate investment trusts", "equity reits", "mortgage reits"],
        "strong": ["real estate", "property", "reit", "rental", "leasing", "mortgage reit"],
        "medium": ["property management", "realty", "commercial real estate"],
        "weak": ["land", "estate"],
    },
    "Transportation": {
        "very_strong": ["airline", "railroad", "shipping", "logistics", "freight"],
        "strong": [
            "transportation", "transport", "rail", "railroad", "airline", "airport",
            "shipping", "cargo", "truck", "logistics", "warehouse", "freight", "marine",
        ],
        "medium": ["distribution center", "ground transport", "air transport", "port"],
        "weak": ["infrastructure", "delivery"],
    },
    "Consumer": {
        "very_strong": ["hospitality services", "food and beverage", "retail", "restaurants"],
        "strong": [
            "retail", "food", "beverage", "tobacco", "apparel", "restaurant", "hotel",
            "hospitality", "leisure", "entertainment", "gaming", "consumer",
        ],
        "medium": ["household products", "travel services", "consumer products", "media"],
        "weak": ["household", "stores", "travel", "content"],
    },
    "Industrials/Materials": {
        "very_strong": ["industrial manufacturing", "aerospace and defense", "mining and mineral products"],
        "strong": [
            "industrial", "machinery", "chemical", "chemicals", "plastic", "rubber",
            "metal", "mining", "mineral", "construction", "aerospace", "defense",
            "equipment", "manufacturing",
        ],
        "medium": ["engineering", "capital goods", "building products", "paper", "packaging"],
        "weak": ["materials", "services", "fabrication"],
    },
    "Public/Education": {
        "very_strong": ["public administration", "educational services"],
        "strong": ["public administration", "government", "education", "school", "university"],
        "medium": ["public sector", "k 12", "higher education"],
        "weak": ["municipal", "state"],
    },
}

BUCKET_WEIGHTS: Dict[str, float] = {
    "very_strong": 5.0,
    "strong": 3.0,
    "medium": 2.0,
    "weak": 1.0,
}


def _clean_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _normalize_label(value: object) -> str:
    text = _clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _join_text_parts(parts: List[object]) -> str:
    return " ".join([_clean_text(p) for p in parts if _clean_text(p)])


def parse_sic(value: object) -> Optional[int]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = re.sub(r"\D", "", str(value))
    if not text:
        return None
    try:
        num = int(text)
    except ValueError:
        return None
    if num < 0 or num > 9999:
        return None
    return num


def load_ff49_sic_ranges(url: str = SIC49_URL) -> pd.DataFrame:
    response = requests.get(url, timeout=30)
    response.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(response.content)) as zf:
        names = zf.namelist()
        if not names:
            raise ValueError("Siccodes49.zip did not contain any files.")
        raw_text = zf.read(names[0]).decode("latin1", errors="replace")

    rows: List[Dict[str, object]] = []
    current: Dict[str, object] = {}

    header_re = re.compile(r"^\s*(\d+)\s+([A-Za-z0-9]+)\s{2,}(.+?)\s*$")
    range_re = re.compile(r"^\s*(\d{4})-(\d{4})\s+(.+?)\s*$")

    for raw_line in raw_text.splitlines():
        line = raw_line.replace("\r", "")

        header_match = header_re.match(line)
        if header_match:
            current = {
                "FF49_Num": int(header_match.group(1)),
                "FF49_Code": header_match.group(2).strip(),
                "FF49_Industry": header_match.group(3).strip(),
            }
            continue

        range_match = range_re.match(line)
        if range_match and current:
            rows.append(
                {
                    **current,
                    "SIC_Start": int(range_match.group(1)),
                    "SIC_End": int(range_match.group(2)),
                    "SIC_Description": range_match.group(3).strip(),
                }
            )

    if not rows:
        raise ValueError("Failed to parse SIC ranges from Ken French SIC49 file.")

    out = pd.DataFrame(rows).sort_values(["SIC_Start", "SIC_End", "FF49_Num"]).reset_index(drop=True)
    return out


def naics_sector_from_code(naics_code: object) -> str:
    cleaned = re.sub(r"\D", "", _clean_text(naics_code))
    if len(cleaned) < 2:
        return "Unknown"
    two = int(cleaned[:2])

    if two == 11:
        return "11 Agriculture, Forestry, Fishing and Hunting"
    if two == 21:
        return "21 Mining, Quarrying, and Oil and Gas Extraction"
    if two == 22:
        return "22 Utilities"
    if two == 23:
        return "23 Construction"
    if two in (31, 32, 33):
        return "31-33 Manufacturing"
    if two == 42:
        return "42 Wholesale Trade"
    if two in (44, 45):
        return "44-45 Retail Trade"
    if two in (48, 49):
        return "48-49 Transportation and Warehousing"
    if two == 51:
        return "51 Information"
    if two == 52:
        return "52 Finance and Insurance"
    if two == 53:
        return "53 Real Estate and Rental and Leasing"
    if two == 54:
        return "54 Professional, Scientific, and Technical Services"
    if two == 55:
        return "55 Management of Companies and Enterprises"
    if two == 56:
        return "56 Administrative and Support and Waste Management and Remediation Services"
    if two == 61:
        return "61 Educational Services"
    if two == 62:
        return "62 Health Care and Social Assistance"
    if two == 71:
        return "71 Arts, Entertainment, and Recreation"
    if two == 72:
        return "72 Accommodation and Food Services"
    if two == 81:
        return "81 Other Services (except Public Administration)"
    if two == 92:
        return "92 Public Administration"
    return "Unknown"


def infer_domain(text: object) -> str:
    t = _normalize_label(text)
    if not t:
        return "Unknown"

    scores: Dict[str, float] = {}
    hits: Dict[str, int] = {}
    for domain, buckets in DOMAIN_RULES.items():
        score = 0.0
        hit_count = 0
        for bucket, weight in BUCKET_WEIGHTS.items():
            for kw in buckets.get(bucket, []):
                if kw in t:
                    score += weight
                    hit_count += 1
        scores[domain] = score
        hits[domain] = hit_count

    best_domain, best_score = max(scores.items(), key=lambda x: x[1])
    if best_score <= 0:
        return "Other"

    # Secondary tie-break by number of matched keywords.
    max_hits = max(hits.values())
    hit_tied = [d for d, h in hits.items() if h == max_hits]
    scored_tied = [d for d, s in scores.items() if s == best_score]
    tied = [d for d in scored_tied if d in hit_tied] or scored_tied

    # Tie-break by stronger lexical specificity for Technology and Healthcare over generic industrial terms.
    priority = [
        "Technology",
        "Healthcare",
        "Financials",
        "Energy",
        "Utilities",
        "Real Estate",
        "Transportation",
        "Consumer",
        "Industrials/Materials",
        "Public/Education",
    ]
    for domain in priority:
        if domain in tied:
            return domain
    return best_domain


def map_sic_to_ff49(base_df: pd.DataFrame, ff49_ranges: pd.DataFrame) -> pd.DataFrame:
    interval_index = pd.IntervalIndex.from_arrays(ff49_ranges["SIC_Start"], ff49_ranges["SIC_End"], closed="both")

    def lookup(sic_num: Optional[int]) -> Dict[str, object]:
        if sic_num is None:
            return {"FF49_Num": pd.NA, "FF49_Code": pd.NA, "FF49_Industry": pd.NA}
        pos = interval_index.get_indexer([sic_num])[0]
        if pos == -1:
            return {"FF49_Num": pd.NA, "FF49_Code": pd.NA, "FF49_Industry": pd.NA}
        row = ff49_ranges.iloc[pos]
        return {
            "FF49_Num": int(row["FF49_Num"]),
            "FF49_Code": row["FF49_Code"],
            "FF49_Industry": row["FF49_Industry"],
        }

    mapped = base_df["SIC_Int"].apply(lookup).apply(pd.Series)
    return pd.concat([base_df, mapped], axis=1)


def build_output(df: pd.DataFrame, ff49_ranges: pd.DataFrame) -> pd.DataFrame:
    required = ["Symbol", "SIC_CODES", "NAICS Code", "FR RBICS Name Sector"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required columns in '{SOURCE_SHEET}': {missing}")

    out = df.copy()
    out["SIC_Int"] = out["SIC_CODES"].apply(parse_sic)
    out = map_sic_to_ff49(out, ff49_ranges)

    out["NAICS_Sector_Label"] = out["NAICS Code"].apply(naics_sector_from_code)
    rbics_text = out.apply(
        lambda r: _join_text_parts(
            [
                r.get("FR RBICS Name Sector"),
                r.get("FR RBICS Name Subsector"),
                r.get("FR RBICS Name Industry"),
                r.get("FR RBICS Name Subindustry"),
            ]
        ),
        axis=1,
    )
    out["FF49_Domain"] = out["FF49_Industry"].apply(infer_domain)
    out["RBICS_Domain"] = rbics_text.apply(infer_domain)
    out["NAICS_Domain"] = out["NAICS_Sector_Label"].apply(infer_domain)

    out["FF49_vs_RBICS_Domain_Match"] = out["FF49_Domain"] == out["RBICS_Domain"]
    out["FF49_vs_NAICS_Domain_Match"] = out["FF49_Domain"] == out["NAICS_Domain"]
    out["RBICS_vs_NAICS_Domain_Match"] = out["RBICS_Domain"] == out["NAICS_Domain"]

    return out


def build_agreement_table(out: pd.DataFrame) -> pd.DataFrame:
    metrics = {
        "Rows": len(out),
        "FF49 coverage (non-null mapped SIC)": out["FF49_Industry"].notna().mean(),
        "FF49 vs RBICS domain agreement": out["FF49_vs_RBICS_Domain_Match"].mean(),
        "FF49 vs NAICS domain agreement": out["FF49_vs_NAICS_Domain_Match"].mean(),
        "RBICS vs NAICS domain agreement": out["RBICS_vs_NAICS_Domain_Match"].mean(),
    }
    rows = []
    for k, v in metrics.items():
        if isinstance(v, (int, float)):
            rows.append({"Metric": k, "Value": round(float(v), 4)})
        else:
            rows.append({"Metric": k, "Value": v})
    return pd.DataFrame(rows)


def build_predictor_table(out: pd.DataFrame) -> pd.DataFrame:
    ff49_rbics = float(out["FF49_vs_RBICS_Domain_Match"].mean())
    ff49_naics = float(out["FF49_vs_NAICS_Domain_Match"].mean())
    rbics_naics = float(out["RBICS_vs_NAICS_Domain_Match"].mean())

    scores = {
        "FF49 (SIC-based)": (ff49_rbics + ff49_naics) / 2.0,
        "RBICS": (ff49_rbics + rbics_naics) / 2.0,
        "NAICS": (ff49_naics + rbics_naics) / 2.0,
    }
    best = max(scores, key=scores.get)

    rows = [
        {"Classification": "FF49 (SIC-based)", "Overlap_Score": round(scores["FF49 (SIC-based)"], 4), "Best_Predictor": best == "FF49 (SIC-based)"},
        {"Classification": "RBICS", "Overlap_Score": round(scores["RBICS"], 4), "Best_Predictor": best == "RBICS"},
        {"Classification": "NAICS", "Overlap_Score": round(scores["NAICS"], 4), "Best_Predictor": best == "NAICS"},
    ]
    return pd.DataFrame(rows)


def write_results(workbook_path: Path, detailed: pd.DataFrame) -> None:
    agreement = build_agreement_table(detailed)
    predictor = build_predictor_table(detailed)

    ff49_vs_rbics_cm = pd.crosstab(
        detailed["FF49_Industry"].fillna("Unmapped SIC"),
        detailed["FR RBICS Name Sector"].fillna("Unknown RBICS"),
        rownames=["FF49_Industry"],
        colnames=["RBICS_Sector"],
    )

    ff49_vs_naics_cm = pd.crosstab(
        detailed["FF49_Industry"].fillna("Unmapped SIC"),
        detailed["NAICS_Sector_Label"].fillna("Unknown NAICS"),
        rownames=["FF49_Industry"],
        colnames=["NAICS_Sector"],
    )

    wb = load_workbook(workbook_path)
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    wb.save(workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        detailed.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False, startrow=0)

        r = 0
        pd.DataFrame({"Section": ["Best Predictor by Overlap"]}).to_excel(
            writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r, header=False
        )
        predictor.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r + 1)

        r = r + 1 + len(predictor) + 2
        pd.DataFrame({"Section": ["Agreement Metrics"]}).to_excel(
            writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r, header=False
        )
        agreement.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r + 1)

        r = r + 1 + len(agreement) + 2
        pd.DataFrame({"Section": ["Confusion Matrix: FF49 vs RBICS Sector"]}).to_excel(
            writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r, header=False
        )
        ff49_vs_rbics_cm.to_excel(writer, sheet_name=SUMMARY_SHEET, startrow=r + 1)

        r = r + 1 + len(ff49_vs_rbics_cm) + 3
        pd.DataFrame({"Section": ["Confusion Matrix: FF49 vs NAICS Sector"]}).to_excel(
            writer, sheet_name=SUMMARY_SHEET, index=False, startrow=r, header=False
        )
        ff49_vs_naics_cm.to_excel(writer, sheet_name=SUMMARY_SHEET, startrow=r + 1)


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    base = pd.read_excel(WORKBOOK_PATH, sheet_name=SOURCE_SHEET)
    ff49_ranges = load_ff49_sic_ranges()
    detailed = build_output(base, ff49_ranges)
    write_results(WORKBOOK_PATH, detailed)

    print(f"Wrote '{OUTPUT_SHEET}' and '{SUMMARY_SHEET}' in {WORKBOOK_PATH}")
    print(f"Rows processed: {len(detailed)}")
    print(f"FF49 mapped rows: {int(detailed['FF49_Industry'].notna().sum())}")


if __name__ == "__main__":
    main()
